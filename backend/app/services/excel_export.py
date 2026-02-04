"""
Excel Export Service

Generates Excel files from evaluation results.
"""

import io
from typing import List
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ..models import MetricResult


def create_excel_report(
    metrics: List[MetricResult],
    filename: str,
    log_file_id: str
) -> bytes:
    """
    Create an Excel report from evaluation metrics.
    
    Returns the Excel file as bytes.
    """
    # Create DataFrame
    data = []
    for metric in metrics:
        data.append({
            "Metric Name": metric.metric_name,
            "Metric Value": str(metric.metric_value),
            "Description": metric.description or ""
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Metrics', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Metrics']
        
        # Style the header row
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 60
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(data) + 1, max_col=3):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Add metadata sheet
        metadata_sheet = workbook.create_sheet(title='Info')
        metadata_sheet['A1'] = 'Log File ID'
        metadata_sheet['B1'] = log_file_id
        metadata_sheet['A2'] = 'Original Filename'
        metadata_sheet['B2'] = filename
        metadata_sheet['A3'] = 'Total Metrics'
        metadata_sheet['B3'] = len(metrics)
        
        for cell in metadata_sheet['A']:
            cell.font = Font(bold=True)
        
        metadata_sheet.column_dimensions['A'].width = 20
        metadata_sheet.column_dimensions['B'].width = 40
    
    output.seek(0)
    return output.getvalue()


def metrics_to_dataframe(metrics: List[MetricResult]) -> pd.DataFrame:
    """Convert metrics to a pandas DataFrame."""
    data = []
    for metric in metrics:
        data.append({
            "Metric Name": metric.metric_name,
            "Metric Value": str(metric.metric_value),
            "Description": metric.description or ""
        })
    
    return pd.DataFrame(data)
